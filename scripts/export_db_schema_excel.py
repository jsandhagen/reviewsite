import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from database import get_db
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

ROOT = os.path.dirname(os.path.dirname(__file__))
OUT_PATH = os.path.join(ROOT, 'database_schema.xlsx')


def get_tables(conn):
    cur = conn.cursor()
    cur.execute("""
        SELECT tablename FROM pg_tables
        WHERE schemaname='public'
        ORDER BY tablename
    """)
    return [r['tablename'] for r in cur.fetchall()]


def table_info(conn, table):
    cur = conn.cursor()
    cur.execute(f"""
        SELECT
            ordinal_position as cid,
            column_name as name,
            data_type as type,
            CASE WHEN is_nullable = 'NO' THEN 1 ELSE 0 END as notnull,
            column_default as dflt_value,
            CASE WHEN column_default LIKE 'nextval%%' THEN 1 ELSE 0 END as pk
        FROM information_schema.columns
        WHERE table_schema = 'public' AND table_name = %s
        ORDER BY ordinal_position
    """, (table,))
    cols = cur.fetchall()
    return [
        {
            'cid': c['cid'],
            'name': c['name'],
            'type': c['type'],
            'notnull': bool(c['notnull']),
            'dflt_value': c['dflt_value'],
            'pk': bool(c['pk']),
        }
        for c in cols
    ]


def foreign_keys(conn, table):
    cur = conn.cursor()
    cur.execute(f"""
        SELECT
            tc.constraint_name as id,
            0 as seq,
            ccu.table_name as table,
            kcu.column_name as "from",
            ccu.column_name as "to",
            rc.update_rule as on_update,
            rc.delete_rule as on_delete,
            '' as match
        FROM information_schema.table_constraints AS tc
        JOIN information_schema.key_column_usage AS kcu
            ON tc.constraint_name = kcu.constraint_name
            AND tc.table_schema = kcu.table_schema
        JOIN information_schema.constraint_column_usage AS ccu
            ON ccu.constraint_name = tc.constraint_name
            AND ccu.table_schema = tc.table_schema
        JOIN information_schema.referential_constraints AS rc
            ON rc.constraint_name = tc.constraint_name
            AND rc.constraint_schema = tc.table_schema
        WHERE tc.constraint_type = 'FOREIGN KEY'
            AND tc.table_schema = 'public'
            AND tc.table_name = %s
    """, (table,))
    rows = cur.fetchall()
    return [
        {
            'id': idx,
            'seq': r['seq'],
            'table': r['table'],
            'from': r['from'],
            'to': r['to'],
            'on_update': r['on_update'],
            'on_delete': r['on_delete'],
            'match': r['match'],
        }
        for idx, r in enumerate(rows)
    ]


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value is not None else ''
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)


def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    with get_db() as conn:
        tables = get_tables(conn)
        # Ensure preferred order if present
        preferred = ['users', 'games', 'user_scores']
        ordered = [t for t in preferred if t in tables] + [t for t in tables if t not in preferred]

        header_font = Font(bold=True)
        center = Alignment(horizontal='center')

        for t in ordered:
            ws = wb.create_sheet(title=t)
            ws.append(['Column', 'Type', 'Not Null', 'Default', 'Primary Key'])
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = center

            for col in table_info(conn, t):
                ws.append([
                    col['name'],
                    col['type'],
                    'YES' if col['notnull'] else 'NO',
                    col['dflt_value'],
                    'YES' if col['pk'] else 'NO',
                ])
            autosize(ws)

        # Relationships sheet
        ws_rel = wb.create_sheet(title='Relationships')
        ws_rel.append(['From Table', 'From Column', 'To Table', 'To Column', 'On Update', 'On Delete'])
        for cell in ws_rel[1]:
            cell.font = header_font
            cell.alignment = center

        for t in ordered:
            for fk in foreign_keys(conn, t):
                ws_rel.append([t, fk['from'], fk['table'], fk['to'], fk['on_update'], fk['on_delete']])
        autosize(ws_rel)

    wb.save(OUT_PATH)
    print(f"Wrote: {OUT_PATH}")


if __name__ == '__main__':
    main()
